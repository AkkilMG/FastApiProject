
import os
import csv
from openpyxl import load_workbook
from fastapi import FastAPI, Request, File, UploadFile
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse

# internal 
from database.access_db import db

app = FastAPI()
templates = Jinja2Templates(directory="templates")


@app.get("/", response_class=HTMLResponse)
def home_table(request: Request):
    try:
        data = db.getAll()
        # if data == False:
        #     return templates.TemplateResponse("form.html", context)
        # else:
        if 1:
            name = data[0]
            usn = data[1]
            sem = data[2]
            marks = data[3]
            gender = data[4]
            total = data[5]
            id = data[6]
            context={
                "request": request,
                "name": name,
                "usn": usn,
                "sem": sem,
                "marks": marks,
                "gender": gender,
                "total": total,
                "no": len(name),
                "id": id
            }
            return templates.TemplateResponse("table.html", context)
    except Exception as e:
        return {"successful": "False"}

@app.get("/api/create")
def create_student(name:str, usn:str, sem:str, gender:str, marks1:str, marks2:str, marks3:str):
    try:
        marks = [int(marks1), int(marks2), int(marks3)]
        db.insert(name=name, usn=usn, sem=int(sem), gender=gender, marks=marks)
        return {"successful": "True"}
    except Exception:
        return {"successful": "False"}

@app.get("/api/add", response_class=HTMLResponse)
def add(request: Request):
    try:
        context={
            "request": request
        }
        return templates.TemplateResponse("form.html", context)
    except Exception as e:
        return {"successful": "False"}


@app.get("/api/delete/{id}")
def delete_student(request: Request, id: str):
    if id is not None:
        try:
            db.delete(id=id)
            return {"successful": "True"}
        except Exception:
            return {"successful": "False"}
    else:
        return {"successful": "False"}


@app.get("/api/update/{id}")
def update_student(id: str, key: str, value:str):
    if id is not None:
        try:
            db.update(id=id, key=key, value=value)
            return {"successful": "True"}
        except Exception:
            return {"successful": "False"}
    else:
        return {"successful": "False"}

@app.get("/api/edit/{id}", response_class=HTMLResponse)
def edit(request: Request, id: str):
    try:
        context={
            "request": request,
            "id":id
        }

        return templates.TemplateResponse("editForm.html", context)
    except Exception as e:
        return {"successful": "False", "Error": e}

@app.post("/uploadfile")
async def create_upload_file(file: UploadFile):
    try:
        with open(f"./Download/{file.filename}", "wb") as f:
            f.write(await file.read())
        if ".csv" in file.filename:
            with open(f'./Download/{file.filename}', 'r') as csv_file:
                csv_reader = csv.DictReader(csv_file)

                # Iterate over the rows in the CSV file
                for row in csv_reader:
                    # Assign the values in the row to variables
                    name = row['name']
                    usn = row['usn']
                    sem = row['sem']
                    gender = row['gender']
                    marks1 = row['marks1']
                    marks2 = row['marks2']
                    marks3 = row['marks3']
                    ch = db.checker("usn", usn)
                    if(ch == False):
                        marks = [int(marks1), int(marks2), int(marks3)]
                        db.insert(name=name, usn=usn, sem=int(sem), gender=gender, marks=marks)
            try:
                os.remove(f"./Download/{file.filename}")
            except Exception:
                pass
            return {"successful": "True"}
        elif ".xlsx" in file.filename:
            workbook = load_workbook(f'./Download/{file.filename}')
            worksheet = workbook.active
            try:
                for row in worksheet.iter_rows():
                    # data = {}
                    name = row[0].value
                    usn = row[1].value
                    sem = row[2].value
                    gender = row[3].value
                    marks1 = row[4].value
                    marks2 = row[5].value
                    marks3 = row[6].value
                    if row[0].value!="name":
                        ch = db.checker("usn", usn)
                        if(ch == False):
                            marks = [int(marks1), int(marks2), int(marks3)]
                            db.insert(name=name, usn=usn, sem=int(sem), gender=gender, marks=marks)
                try:
                    os.remove(f"./Download/{file.filename}")
                except Exception:
                    pass
                return {"successful": "True"}
            except Exception as e:
                try:
                    os.remove(f"./Download/{file.filename}")
                except Exception:
                    pass
                return {"successful": "False", "Error": e}
    except Exception as e:
        try:
            os.remove(f"./Download/{file.filename}")
        except Exception:
            pass
        return {"successful": "True", "Error": e}


@app.get("/api/semHighest")
def add_entry(request: Request, sem:str):
    try:
        all = db.getAll()
        sem = int(sem)
        max = 0
        for i in range(0, len(all[2])):
            if int(all[2][i]) == sem:
                if i == 0:
                    max = (all[5][0]/3)
                    name = all[0][0]
                    usn = all[1][0]
                    sem = all[2][0]
                    gender = all[4][0]
                if max < all[5][i]:
                    max = all[5][i]
                    name = all[0][i]
                    usn = all[1][i]
                    sem = all[2][i]
                    gender = all[4][i]
        if max != 0:
            return {"successful": "True", "name": name, "usn": usn, "sem": sem, "gender": gender}
        else:
            return {"successful": "False", "Error": "No data available"}
    except Exception as e:
        return {"successful": "False", "Error": e}

@app.get("/api/genderHighest")
def genderHighest(request: Request, sem: str):
    try:
        all = db.getAll()
        sem = int(sem)
        male = 0
        female = 0
        maleNo = 0
        femaleNo = 0
        try:
            for i in range(0, len(all[4])):
                if (all[2][i] == sem):
                    if all[4][i] == "male":
                        male+=all[3][i][0]
                        maleNo+=1
                    elif all[4][i] == "female":
                        female+=all[3][i][0]
                        femaleNo+=1
            
        except Exception as e:
            return {"successful": "False", "Error": e}
        if (male==0) and (female==0):
            return {"successful": "False", "Erro": "No data available."}
        try:
            male = male/maleNo
        except Exception:
            pass
        try:
            female = female/femaleNo
        except Exception:
            pass
        if male > female:
            r = "male"
        else:
            r = "female"

        return {"successful": "True", "gender": r}
    except Exception as e:
        return {"successful": "False", "Error": e}

@app.get("/api/averageHighest")
def add_entry(request: Request):
    try:
        all = db.getAll()
        total = []
        
        for i in range(1,4):
            try:
                mark = 0
                no = 0
                for j in range(0, len(all[2])):
                    if all[2][j] == i:
                        mark+=int(all[5][j]/3)
                        no+=1
                if no == 0:
                    total.append(0)
                else:
                    total.append(mark/no)     
            except Exception as e:
                return {"successful": "False", "Error": e}
        
        for i in range(0, len(total)):
            if i == 0:
                max = total[0]
            if max < total[i]:
                max = total[i]
                nm = i
        return {"successful": "True", "sem": nm+1}
    except Exception as e:
        return {"successful": "False", "Error": e}